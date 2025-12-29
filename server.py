"""
Flask server to handle contact form submissions and save to Excel & send email
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
CORS(app)  # Enable CORS for frontend requests

# Path to Excel file
EXCEL_FILE = Path('docs/contact_submissions.xlsx')

# Email configuration
EMAIL_CONFIG_FILE = 'email_config.txt'
RECIPIENT_EMAILS = ['gregory.rhoney@fora.travel', 'gregory.rhoney@gmail.com']

def load_email_config():
    """Load email configuration from file"""
    if not os.path.exists(EMAIL_CONFIG_FILE):
        return None, None
    
    try:
        with open(EMAIL_CONFIG_FILE, 'r') as f:
            lines = f.readlines()
            email = None
            password = None
            for line in lines:
                line = line.strip()
                if line.startswith('EMAIL='):
                    email = line.split('=', 1)[1]
                elif line.startswith('APP_PASSWORD='):
                    password = line.split('=', 1)[1]
            return email, password
    except Exception as e:
        print(f"Error reading email config: {e}")
        return None, None

def send_email_notification(form_data):
    """Send email notification via SMTP"""
    sender_email, app_password = load_email_config()
    
    if not sender_email or not app_password:
        print("\n⚠️  WARNING: Email not configured!")
        print(f"   Create {EMAIL_CONFIG_FILE} with your Gmail credentials.")
        print("   See EMAIL_SETUP.md for instructions.\n")
        return False
    
    try:
        # Create email content
        subject = f"New SEAL Client Inquiry - {form_data.get('name', 'Unknown')}"
        
        body = f"""
New Contact Form Submission
============================

Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Contact Information:
-------------------
Name: {form_data.get('name', 'N/A')}
Email: {form_data.get('email', 'N/A')}
Phone: {form_data.get('phone', 'N/A')}

Trip Details:
------------
Group Type: {form_data.get('group-type', 'N/A')}
Preferred Dates: {form_data.get('dates', 'N/A')}
Budget Range: {form_data.get('budget', 'N/A')}

Message:
--------
{form_data.get('message', 'N/A')}

---
This submission was automatically sent from the SEAL website contact form.
        """
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(RECIPIENT_EMAILS)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        # Send email via Gmail SMTP
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender_email, app_password)
            server.send_message(msg)
        
        print(f"✓ Email sent to: {', '.join(RECIPIENT_EMAILS)}")
        return True
        
    except Exception as error:
        print(f"Error sending email: {error}")
        return False

def init_excel_file():
    """Initialize Excel file with headers if it doesn't exist"""
    if not EXCEL_FILE.exists():
        EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Contact Submissions"
        
        # Add headers
        headers = ['Timestamp', 'Name', 'Email', 'Phone', 'Group Type', 
                   'Preferred Dates', 'Budget Range', 'Message']
        ws.append(headers)
        
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

@app.route('/submit-contact', methods=['POST'])
def submit_contact():
    """Handle contact form submission"""
    try:
        # Get form data
        data = request.form
        
        # Initialize Excel file if needed
        init_excel_file()
        
        # Load existing workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Prepare row data
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            timestamp,
            data.get('name', ''),
            data.get('email', ''),
            data.get('phone', ''),
            data.get('group-type', ''),
            data.get('dates', ''),
            data.get('budget', ''),
            data.get('message', '')
        ]
        
        # Append data
        ws.append(row_data)
        
        # Save workbook
        wb.save(EXCEL_FILE)
        
        print(f"✓ Excel: {data.get('name')} - {data.get('email')}")
        
        # Try to send email notification
        email_success = send_email_notification(data)
        
        # Return success response
        return jsonify({
            'status': 'success',
            'message': 'Thank you! Your request has been submitted successfully.',
            'saved_to_excel': True,
            'email_sent': email_success
        }), 200
        
    except Exception as e:
        print(f"Error saving submission: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'There was an error submitting your request. Please try again.'
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'}), 200

if __name__ == '__main__':
    print("\n" + "="*60)
    print("Starting SEAL Contact Form Server")
    print("="*60)
    print(f"📊 Excel: {EXCEL_FILE.absolute()}")
    
    # Check email setup
    sender_email, app_password = load_email_config()
    if sender_email and app_password:
        print(f"📧 Email: Enabled")
        print(f"   From: {sender_email}")
        print(f"   To: {', '.join(RECIPIENT_EMAILS)}")
    else:
        print(f"📧 Email: Not configured (add {EMAIL_CONFIG_FILE})")
        print(f"   See EMAIL_SETUP.md for setup instructions")
    
    print("="*60 + "\n")
    
    init_excel_file()
    app.run(host='localhost', port=5000, debug=True)
